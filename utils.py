import os
import re
import threading

# ============ CONFIGURATION OCR ============
OCR_AVAILABLE = False

try:
    from PIL import Image, ImageEnhance
    import pytesseract
    
    if os.name == 'nt':
        possible_paths = [
            r'C:\Program Files\Tesseract-OCR\tesseract.exe',
            r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
        ]
        for path in possible_paths:
            if os.path.exists(path):
                pytesseract.pytesseract.tesseract_cmd = path
                break
    
    OCR_AVAILABLE = True
    print("✅ OCR disponible")
except Exception as e:
    print(f"⚠️ OCR non disponible: {e}")
    OCR_AVAILABLE = False

def process_excel_file(file_path):
    """Traite un fichier Excel"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).lower() if cell.value else "")
        
        name_col = None
        firstname_col = None
        table_col = None
        
        for i, header in enumerate(headers):
            if 'nom' in header or 'name' in header:
                name_col = i
            if 'prénom' in header or 'prenom' in header or 'first' in header:
                firstname_col = i
            if 'table' in header:
                table_col = i
        
        guests = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if name_col is not None and row[name_col]:
                full_name = str(row[name_col])
                parts = full_name.split()
                
                if firstname_col is not None and row[firstname_col]:
                    first_name = str(row[firstname_col]).upper()
                    last_name = full_name.upper()
                else:
                    first_name = parts[0].upper() if parts else ""
                    last_name = " ".join(parts[1:]).upper() if len(parts) > 1 else ""
                
                if table_col is not None and row[table_col]:
                    table = str(row[table_col])
                else:
                    table = "À assigner"
                
                if first_name and last_name:
                    guests.append({
                        'first_name': first_name.strip(),
                        'last_name': last_name.strip(),
                        'table_number': table
                    })
        
        return guests
    
    except Exception as e:
        raise Exception(f"Erreur fichier: {str(e)}")

def process_image_file(file_path):
    """Analyse une photo de liste et extrait les invités"""
    if not OCR_AVAILABLE:
        raise Exception("L'OCR n'est pas disponible")
    
    try:
        image = Image.open(file_path)
        image = image.convert('L')
        
        enhancer = ImageEnhance.Contrast(image)
        image = enhancer.enhance(2.0)
        
        # OCR avec différentes configurations
        all_text = ""
        try:
            text = pytesseract.image_to_string(image, lang='fra')
            all_text += text + "\n"
        except:
            pass
        
        try:
            text = pytesseract.image_to_string(image)
            all_text += text + "\n"
        except:
            pass
        
        print(f"=== TEXTE OCR EXTRAIT ===")
        print(all_text)
        print(f"=== FIN TEXTE ===")
        
        guests = smart_parse_text(all_text)
        print(f"=== INVITÉS TROUVÉS : {len(guests)} ===")
        for g in guests:
            print(f"  {g['first_name']} {g['last_name']} - Table {g['table_number']}")
        
        return guests
    
    except Exception as e:
        raise Exception(f"Erreur analyse photo: {str(e)}")

def smart_parse_text(text):
    """Parse agressif - Détecte les noms et tables dans n'importe quel format"""
    guests = []
    seen_names = set()
    
    # Nettoyer le texte
    text = text.replace('|', ' ').replace('\t', ' ').replace('  ', ' ')
    text = text.replace('→', ' ').replace('->', ' ').replace('—', ' ')
    text = text.replace('_', ' ').replace('.', ' ')
    
    lines = text.strip().split('\n')
    
    for line in lines:
        line = line.strip()
        if not line or len(line) < 3:
            continue
        
        # Ignorer les lignes d'en-tête évidentes
        lower = line.lower()
        if any(w in lower for w in ['nom', 'prénom', 'prenom', 'liste', 'invité', 'invite', 'page', 'total', 'tableau']):
            if len(line.split()) <= 4:
                continue
        
        words = line.split()
        
        # Trouver les mots en MAJUSCULES (noms de famille)
        uppercase_words = []
        for w in words:
            # Nettoyer le mot des ponctuations
            clean_w = w.strip(':,;-()[]{}')
            if clean_w.isupper() and len(clean_w) >= 2 and not clean_w.isdigit():
                uppercase_words.append(clean_w)
        
        # Trouver les mots avec première lettre majuscule (prénoms)
        capitalized_words = []
        for w in words:
            clean_w = w.strip(':,;-()[]{}')
            if (clean_w[0].isupper() if clean_w else False) and not clean_w.isupper() and len(clean_w) >= 2:
                capitalized_words.append(clean_w)
        
        # Trouver les numéros de table
        numbers = []
        for w in words:
            clean_w = w.strip(':,;-()[]{}')
            if clean_w.isdigit():
                numbers.append(clean_w)
        
        # Chercher le mot "Table" ou "Tbl"
        table_name = None
        for i, w in enumerate(words):
            clean_w = w.lower().strip(':,;')
            if clean_w in ['table', 'tbl', 'tab', 't']:
                if i + 1 < len(words):
                    table_name = words[i + 1].strip(':,;-()[]{}')
                break
        
        # ============ DÉTECTION DES NOMS ============
        
        # Format 1 : "DUPONT Marie Table 8" ou "DUPONT Marie 8"
        if len(uppercase_words) >= 1 and len(capitalized_words) >= 1:
            last_name = uppercase_words[0]
            first_name = capitalized_words[0]
            
            if table_name:
                table = table_name
            elif numbers:
                table = numbers[0]
            else:
                table = "À assigner"
            
            key = f"{last_name}_{first_name}"
            if key not in seen_names:
                seen_names.add(key)
                guests.append({
                    'first_name': first_name.upper(),
                    'last_name': last_name.upper(),
                    'table_number': table.upper()
                })
            continue
        
        # Format 2 : "Marie DUPONT Table 8" (prénom d'abord)
        if len(uppercase_words) >= 1 and len(words) >= 2:
            last_name = uppercase_words[0]
            
            # Chercher le prénom (mot avant le nom)
            first_name = None
            for i, w in enumerate(words):
                clean_w = w.strip(':,;-()[]{}')
                if clean_w == last_name and i > 0:
                    first_name = words[i-1].strip(':,;-()[]{}')
                    break
            
            # Si pas trouvé avant, chercher après
            if not first_name:
                for i, w in enumerate(words):
                    clean_w = w.strip(':,;-()[]{}')
                    if clean_w == last_name and i < len(words) - 1:
                        first_name = words[i+1].strip(':,;-()[]{}')
                        break
            
            if first_name and len(first_name) >= 2 and not first_name.isdigit():
                if table_name:
                    table = table_name
                elif numbers:
                    table = numbers[0]
                else:
                    table = "À assigner"
                
                key = f"{last_name}_{first_name}"
                if key not in seen_names:
                    seen_names.add(key)
                    guests.append({
                        'first_name': first_name.upper(),
                        'last_name': last_name.upper(),
                        'table_number': table.upper()
                    })
    
    # Si AUCUN invité trouvé avec les patterns, essayer ligne par ligne simple
    if len(guests) == 0:
        for line in lines:
            line = line.strip()
            if not line or len(line) < 3:
                continue
            
            words = line.split()
            if len(words) < 2:
                continue
            
            # Chercher un numéro de table
            table = "À assigner"
            name_words = []
            
            for w in words:
                clean_w = w.strip(':,;-()[]{}')
                if clean_w.isdigit():
                    table = clean_w
                elif len(clean_w) >= 2:
                    name_words.append(clean_w)
            
            if len(name_words) >= 2:
                # Premier mot = prénom, deuxième = nom (ou inversement)
                if name_words[0].isupper() and not name_words[1].isupper():
                    last_name = name_words[0]
                    first_name = name_words[1]
                elif name_words[1].isupper() and not name_words[0].isupper():
                    last_name = name_words[1]
                    first_name = name_words[0]
                else:
                    first_name = name_words[0]
                    last_name = name_words[1]
                
                key = f"{last_name}_{first_name}"
                if key not in seen_names:
                    seen_names.add(key)
                    guests.append({
                        'first_name': first_name.upper(),
                        'last_name': last_name.upper(),
                        'table_number': table.upper()
                    })
    
    return guests

def announce_table(table_number):
    print(f"Annonce vocale: Table {table_number}")

def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls', 'jpg', 'jpeg', 'png', 'bmp', 'tiff', 'webp'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def is_image_file(filename):
    IMAGE_EXTENSIONS = {'jpg', 'jpeg', 'png', 'bmp', 'tiff', 'webp'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in IMAGE_EXTENSIONS